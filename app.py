"""
Bank Statement -> CSV/Excel Extractor (Accountant-Grade)
========================================================

Supported banks: Halkbank (V1 + V2 layouts), Akbank, Ziraat (vector + scanned/OCR).
Unknown banks: generic fallback parser with auto-detected row patterns.

Output columns: TARIH | Saat | TUTAR | Bakiye | ACIKLAMA | DEKONT

DEPLOYMENT
----------
This app is deployment-friendly. It runs on:

  * Streamlit Community Cloud — push the repo with `requirements.txt`
    and `packages.txt` (lists tesseract-ocr, tesseract-ocr-tur, poppler-utils).
    No code changes needed.
  * Linux server / Docker — `apt install tesseract-ocr tesseract-ocr-tur
    poppler-utils` then `pip install -r requirements.txt`.
  * macOS local — `brew install tesseract tesseract-lang poppler`
    then `pip install -r requirements.txt`.
  * Windows local — install tesseract from
    https://github.com/UB-Mannheim/tesseract/wiki (default path
    C:\\Program Files\\Tesseract-OCR) and poppler from
    https://github.com/oschwartz10612/poppler-windows/releases (default
    path C:\\Program Files\\poppler\\Library\\bin). Override either path
    via env vars TESSERACT_CMD and POPPLER_PATH if installed elsewhere.

If the OCR libraries / system binaries are missing, the app still works
for vector PDFs (Halkbank, Akbank, vector-mode Ziraat). Only scanned
PDFs need OCR.
"""

from __future__ import annotations

import io
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date as _date
from decimal import Decimal, InvalidOperation
from typing import Callable, Optional

import pandas as pd
import pdfplumber
import streamlit as st

# -----------------------------------------------------------------------------
# Cross-platform OCR setup
# -----------------------------------------------------------------------------
# pytesseract + pdf2image are optional (only needed for scanned PDFs).
# On Linux/macOS/Streamlit-Cloud, the tesseract & poppler binaries live on the
# system PATH (installed via apt/brew/packages.txt) and `pdf2image` accepts
# `poppler_path=None` to mean "use PATH".
# On Windows the binaries usually aren't on PATH, so we look in default install
# locations and let env vars override.

OCR_AVAILABLE: bool = False
OCR_IMPORT_ERROR: Optional[str] = None
POPPLER_PATH: Optional[str] = None  # None = use system PATH

try:
    import pytesseract  # type: ignore
    from pdf2image import convert_from_bytes  # type: ignore

    if sys.platform == 'win32':
        # Tesseract: env var TESSERACT_CMD wins; otherwise default install path.
        _tess_cmd = os.environ.get('TESSERACT_CMD') or \
            r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        if os.path.exists(_tess_cmd):
            pytesseract.pytesseract.tesseract_cmd = _tess_cmd
        # Poppler: env var POPPLER_PATH wins; otherwise default install path.
        _poppler = os.environ.get('POPPLER_PATH') or \
            r'C:\Program Files\poppler\Library\bin'
        if os.path.exists(_poppler):
            POPPLER_PATH = _poppler

    # Confirm the tesseract binary is actually callable.
    pytesseract.get_tesseract_version()
    OCR_AVAILABLE = True
except ImportError as _exc:
    OCR_IMPORT_ERROR = (
        f"Python OCR libraries not installed: {_exc}. "
        "Install with: pip install pytesseract pdf2image"
    )
except Exception as _exc:
    OCR_IMPORT_ERROR = (
        f"OCR libraries imported but tesseract binary not found: {_exc}. "
        "Install tesseract via your system package manager (apt/brew/installer) "
        "and ensure the 'tur' (Turkish) language pack is included."
    )


# =============================================================================
# 1. Number parsing
# =============================================================================

def parse_tr_decimal(s: str) -> Decimal:
    s = s.strip()
    if not s:
        raise InvalidOperation("empty number string")
    neg = s.startswith('-')
    if neg:
        s = s[1:]
    s = s.replace('.', '').replace(',', '.')
    d = Decimal(s)
    return -d if neg else d


def format_tr_decimal(d: Optional[Decimal]) -> str:
    if d is None:
        return ''
    s = f"{d:,.2f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


# =============================================================================
# 2. Result containers
# =============================================================================

@dataclass
class TransactionRow:
    date: str
    time: str = ''
    amount: Decimal = Decimal(0)
    balance: Decimal = Decimal(0)
    description: str = ''
    receipt: str = ''


@dataclass
class StatementMetadata:
    bank: str = ''
    customer_no: str = ''
    customer_name: str = ''
    account_no: str = ''
    iban: str = ''
    branch: str = ''
    currency: str = ''
    period_start: str = ''
    period_end: str = ''
    stated_balance: Optional[Decimal] = None


@dataclass
class BalanceIssue:
    row_index: int
    date: str
    expected: Decimal
    actual: Decimal
    diff: Decimal


@dataclass
class StatementResult:
    source_filename: str
    metadata: StatementMetadata
    rows: list[TransactionRow] = field(default_factory=list)
    balance_issues: list[BalanceIssue] = field(default_factory=list)
    final_balance_match: Optional[bool] = None
    parser_warnings: list[str] = field(default_factory=list)  # actionable
    parser_info: list[str] = field(default_factory=list)      # informational


def _fix_ocr_date(s: str) -> str:
    """Validate DD.MM.YYYY (or DD-MM-YYYY) and try common OCR digit
    substitutions if invalid.

    OCR commonly misreads digits in similar-looking fonts:
      5 <-> 3 (curve at top of 5 mistaken for 3, or vice versa)
      8 <-> 0 / 6 <-> 0 / 9 <-> 4 / 1 <-> 7

    Returns a normalized 'DD.MM.YYYY' string, or '' if unfixable.
    """
    if not s:
        return ''
    m = re.match(r'^\s*(\d{2})[.\-/\s](\d{2})[.\-/\s](\d{4})\s*$', s)
    if not m:
        return ''
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))

    def _valid(dd: int, mm: int, yy: int) -> bool:
        try:
            _date(yy, mm, dd)
            return True
        except (ValueError, TypeError):
            return False

    if _valid(d, mo, y):
        return f'{d:02d}.{mo:02d}.{y}'

    # OCR digit confusion table — only substitutions known to occur.
    confusions = {
        '0': ['8', '6', '9'], '1': ['7', '4'], '2': ['7'],
        '3': ['5', '8'],      '4': ['9', '1'], '5': ['3', '6', '8'],
        '6': ['0', '5', '8'], '7': ['1', '2'], '8': ['0', '3', '5', '6'],
        '9': ['4', '0'],
    }
    # Try fixing the day digits (most common: '50' read for '30').
    digits = f'{d:02d}'
    for pos, ch in enumerate(digits):
        for repl in confusions.get(ch, []):
            cand_d = int(digits[:pos] + repl + digits[pos + 1:])
            if _valid(cand_d, mo, y):
                return f'{cand_d:02d}.{mo:02d}.{y}'
    # Then try the month digits.
    digits = f'{mo:02d}'
    for pos, ch in enumerate(digits):
        for repl in confusions.get(ch, []):
            cand_mo = int(digits[:pos] + repl + digits[pos + 1:])
            if _valid(d, cand_mo, y):
                return f'{d:02d}.{cand_mo:02d}.{y}'
    return ''


# =============================================================================
# 3. Bank detection
# =============================================================================

# Turkish IBAN bank codes — positions 5-9 of the 26-char TR IBAN.
# Only includes banks where statements have been observed; extending this
# table is the primary way to add support for a new bank.
_IBAN_BANK_CODES: dict[str, str] = {
    '00010': 'ZIRAAT',          # Türkiye Cumhuriyeti Ziraat Bankası
    '00012': 'HALKBANK',        # Türkiye Halk Bankası
    '00015': 'VAKIFBANK',       # Türkiye Vakıflar Bankası
    '00032': 'TEB',             # Türk Ekonomi Bankası
    '00046': 'AKBANK',
    '00059': 'SEKERBANK',       # Şekerbank
    '00062': 'GARANTI',         # Garanti BBVA / Türkiye Garanti Bankası
    '00064': 'ISBANK',          # Türkiye İş Bankası
    '00067': 'YAPIKREDI',       # Yapı ve Kredi Bankası
    '00099': 'INGBANK',         # ING Bank
    '00103': 'FIBABANKA',       # Fibabanka
    '00111': 'QNB',             # QNB Finansbank
    '00123': 'HSBC',            # HSBC Bank
    '00124': 'ALTERNATIF',      # Alternatifbank
    '00125': 'BURGAN',          # Burgan Bank
    '00134': 'DENIZBANK',
    '00146': 'ODEABANK',        # Odeabank
    '00203': 'ALBARAKA',        # Albaraka Türk
    '00205': 'KUVEYTTURK',      # Kuveyt Türk Katılım Bankası
    '00206': 'TURKIYEFINANS',   # Türkiye Finans Katılım Bankası
    '00209': 'VAKIFKATILIM',    # Vakıf Katılım Bankası
    '00210': 'ZIRAATKATILIM',   # Ziraat Katılım Bankası
}


def _extract_iban_bank_code(text: str) -> Optional[str]:
    """Find the customer's IBAN in the metadata header (first ~2000 chars,
    before the transaction table starts) and return the 5-digit bank code.

    Handles several real-world label variants:
      'IBAN: TR84 0001 2 ...'
      'IBAN :TR84000120012...'
      'IBAN/Hesap No TR31002060...'   (slash-joined label)
      'IBAN/Hesap No : TR82000670...'
    """
    head = text[:2000]
    # Strip whitespace inside IBAN values (some banks render with spaces every 4 chars).
    head_compact = re.sub(r'(TR\d{2})\s+', r'\1', head)
    head_compact = re.sub(r'(\d{4})\s+(\d)', r'\1\2', head_compact)

    # Strict: explicit IBAN label (any variant), then TR + check + 5-digit bank code.
    m = re.search(
        r'IBAN(?:\s*/\s*[\w\s]+?)?\s*[:\s]\s*TR\d{2}(\d{5})',
        head_compact,
    )
    if m:
        return m.group(1)
    # Looser: any TR-IBAN-like token in header.
    m = re.search(r'\bTR\d{2}(\d{5})\d', head_compact)
    if m:
        return m.group(1)
    return None


def detect_bank(full_text: str) -> Optional[str]:
    """Detect the issuing bank using a multi-signal approach:
      1. Customer's IBAN bank code (most reliable — unique per bank,
         lives in the metadata header, immune to transaction-text noise).
      2. Scoring on header + body text signals (fallback when IBAN is
         missing or masked too aggressively).
    """
    # Strongest signal first: IBAN bank code.
    code = _extract_iban_bank_code(full_text)
    if code and code in _IBAN_BANK_CODES:
        return _IBAN_BANK_CODES[code]

    # Fallback: scoring on text signals.
    upper = full_text.upper()
    head = upper[:500]
    body = upper[:8000]
    scores: dict[str, int] = {}

    def add(bank: str, points: int) -> None:
        scores[bank] = scores.get(bank, 0) + points

    # Halkbank
    if 'HALKBANK' in head:
        add('HALKBANK', 10)
    if 'TÜRKIYE HALK BANKASI' in body or 'TURKIYE HALK BANKASI' in body:
        add('HALKBANK', 8)
    if 'HALKBANK.COM.TR' in body:
        add('HALKBANK', 8)
    if 'HALKBANK DIALOG' in body:
        add('HALKBANK', 6)
    if 'MÜŞTERI BILGILERINIZ' in body or 'MÜŞTERİ BİLGİLERİNİZ' in body:
        add('HALKBANK', 4)

    # Akbank
    if 'AKBANK' in head:
        add('AKBANK', 10)
    if 'AKBANK T.A.S' in body or 'AKBANK T.A.Ş' in body:
        add('AKBANK', 8)
    if 'AKPOS' in body:
        add('AKBANK', 4)

    # Ziraat — only count strong header signals, NOT mid-text Ziraat refs
    # which appear constantly in transaction descriptions of OTHER banks
    # (transfers to/from Ziraat accounts).
    if 'ZIRAAT BANKASI' in head or 'ZİRAAT BANKASI' in head:
        add('ZIRAAT', 10)
    if 'T.C. ZIRAAT' in head or 'T.C. ZİRAAT' in head:
        add('ZIRAAT', 10)
    if 'ZIRAATBANK.COM.TR' in body:
        add('ZIRAAT', 8)

    # Garanti BBVA
    if 'GARANTI BBVA' in body or 'GARANTİ BBVA' in body:
        add('GARANTI', 10)
    if 'GARANTI BANKASI' in head or 'GARANTİ BANKASI' in head:
        add('GARANTI', 10)

    # Yapı Kredi
    if 'YAPI VE KREDI' in head or 'YAPI KREDI' in head or 'YAPIKREDI' in head:
        add('YAPIKREDI', 10)

    # İş Bankası
    if 'İŞ BANKASI' in head or 'IS BANKASI' in head or 'TÜRKİYE İŞ BANKASI' in head:
        add('ISBANK', 10)

    # Vakıfbank
    if 'VAKIFBANK' in head or 'VAKIFLAR BANKASI' in head:
        add('VAKIFBANK', 10)

    if scores:
        return max(scores.items(), key=lambda kv: kv[1])[0]
    return None


# =============================================================================
# 4. PDF text extraction with OCR fallback
# =============================================================================

def _is_pdf_scanned(pdf_bytes: bytes) -> bool:
    """A PDF is considered scanned if pdfplumber finds zero text characters
    across all pages (only images / no embedded text layer)."""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_chars = sum(len(p.chars) for p in pdf.pages)
        return total_chars == 0
    except Exception:
        return False


def extract_pdf_text(pdf_bytes: bytes) -> str:
    """Vector-text extraction. Returns '' if the PDF is image-only."""
    parts: list[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or '')
    return '\n'.join(parts)


def _ocr_pages(pdf_bytes: bytes, dpi: int = 400, lang: str = 'tur') -> list[dict]:
    """OCR every page and return list of {text, words, image_size}.
    `words` is the per-token positional data (used by Ziraat OCR parser)."""
    if not OCR_AVAILABLE:
        raise RuntimeError(
            "OCR libraries not installed. To process scanned PDFs install:\n"
            "  pip install pytesseract pdf2image\n"
            "  apt-get install tesseract-ocr tesseract-ocr-tur poppler-utils\n"
            f"Original import error: {OCR_IMPORT_ERROR}"
        )
    images = convert_from_bytes(pdf_bytes, dpi=dpi, poppler_path=POPPLER_PATH)
    results = []
    for img in images:
        data = pytesseract.image_to_data(
            img, lang=lang, output_type=pytesseract.Output.DICT, config='--psm 6'
        )
        # Group OCR words into lines by block/par/line, sorted by y then x
        lines: dict = {}
        for i, txt in enumerate(data['text']):
            if not txt.strip():
                continue
            key = (data['block_num'][i], data['par_num'][i], data['line_num'][i])
            lines.setdefault(key, []).append((data['left'][i], data['top'][i], txt))
        sorted_lines = sorted(
            lines.items(), key=lambda kv: min(t for _, t, _ in kv[1])
        )
        line_objs = []
        text_lines = []
        for _key, words in sorted_lines:
            words.sort()  # by left x
            tokens = [t for _, _, t in words]
            line_objs.append({'tokens': tokens, 'top': min(t for _, t, _ in words)})
            text_lines.append(' '.join(tokens))
        results.append({
            'text': '\n'.join(text_lines),
            'lines': line_objs,
            'size': img.size,
        })
    return results


# =============================================================================
# 5. Halkbank parser  (unchanged from original)
# =============================================================================

_HB_SKIP_CONT_PREFIXES = (
    'Müşteri', 'Hesap', 'TCKN', 'IBAN', 'Şube', 'Döviz', 'Üretim', 'Dönemi',
    'Bakiye Bilgi', 'Bloke', 'Kullanıl', 'Toplam Kredi',
    'Türkiye Halk', 'yerine kullanıl', 'Uyuşmazlık',
    'MÜŞTERİ', 'Dönem',
)
_HB_SKIP_EXACT = {'HESAP ÖZETİ'}

_HB_NUM = r'-?[\d.]+,\d{2}'
# Allow ZERO whitespace between the date and a negative amount: some Halkbank
# layouts render '30-04-2025-40.312,86' (the '-' of the amount glued to the
# date). We use \s* between date and amount; the regex still disambiguates
# correctly because \d{4} is non-greedy w.r.t. extra digits and the amount
# starts with -? followed by [\d.]+,\d{2}.
_HB_ROW_V1 = re.compile(rf'^(\d{{2}}-\d{{2}}-\d{{4}})\s*({_HB_NUM})\s+({_HB_NUM})\s+(.*)$')

_HB_NUM_UNSIGNED = r'[\d.]+,\d{2}'
_HB_ROW_V2 = re.compile(
    rf'^(\d{{2}}\.\d{{2}}\.\d{{4}})\s+'
    rf'\d{{2}}\.\d{{2}}\.\d{{4}}\s+'
    rf'({_HB_NUM_UNSIGNED})\s+([+-])\s+'
    rf'({_HB_NUM_UNSIGNED})\s+([+-])\s+'
    rf'(.*)$'
)

_HB_DEKONT_RE = re.compile(r'/(\d{10,})\s*$')


def _hb_extract_dekont(description: str) -> str:
    m = _HB_DEKONT_RE.search(description)
    return m.group(1) if m else ''


def _hb_strip_right_column(line: str) -> str:
    right_anchors = (
        r'\s+(?:Üretim\s+Zamanı|Dönemi|Hesap\s+Bakiyesi|Bloke\s+Bakiyesi|'
        r'Kullanılabilir(?:\s+\w+)*|Toplam\s+Kredi|Bakiye\s+Bilgileriniz|'
        r'Hesap\s+Özeti)\s*:'
    )
    return re.split(right_anchors, line, maxsplit=1)[0].rstrip()


def _hb_extract_customer_name(text: str) -> str:
    lines = text.split('\n')
    for line in lines:
        m = re.search(r'Müşteri Adı\s*/\s*Ünvanı\s*:\s*(.+)', line)
        if m:
            name = m.group(1).strip()
            if name:
                return name
    label_idx = next(
        (i for i, l in enumerate(lines) if 'Müşteri Adı / Ünvanı' in l), None)
    if label_idx is None:
        return ''
    next_label_re = re.compile(r'^(TCKN|Hesap|Bakiye|Şube|Döviz|IBAN|Bloke|Toplam)\b')
    parts: list[str] = []
    if label_idx > 0:
        cand = _hb_strip_right_column(lines[label_idx - 1]).strip()
        if cand and not cand.startswith(('Müşteri', 'Hesap', 'TCKN', 'IBAN',
                                         'Şube', 'Döviz', 'Bakiye', 'Bloke',
                                         'Kullanıl', 'Toplam', 'Üretim', 'Dönemi',
                                         'MÜŞTERİ', 'Dönem')):
            parts.append(cand)
    for j in range(label_idx + 1, min(label_idx + 5, len(lines))):
        cand = _hb_strip_right_column(lines[j]).strip()
        if not cand:
            continue
        if next_label_re.match(cand):
            break
        parts.append(cand)
    return ' '.join(parts).strip()


def _hb_extract_metadata(pdf_bytes: bytes) -> StatementMetadata:
    md = StatementMetadata(bank='HALKBANK')
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
    except Exception:
        return md
    cleaned_lines = [_hb_strip_right_column(l) for l in text.split('\n')]
    cleaned = '\n'.join(cleaned_lines)

    def grab(pattern: str, source: str = cleaned) -> Optional[str]:
        m = re.search(pattern, source)
        return m.group(1).strip() if m else None

    md.customer_no = grab(r'Müşteri Numarası\s*:\s*(\S+)') or grab(r'Müşteri No\s*:\s*(\S+)') or ''
    md.customer_name = _hb_extract_customer_name(text)
    md.account_no = grab(r'Hesap No\s*:\s*(\S+)') or ''
    md.iban = grab(r'IBAN\s*:\s*(\S+)') or ''
    md.branch = grab(r'Şube Kodu / Adı\s*:\s*([^\n]+)') or ''
    if not md.branch:
        code = grab(r'Şube Kodu\s*:\s*(\S+)') or ''
        name = grab(r'Şube Adı\s*:\s*([^\n]+)') or ''
        if code or name:
            md.branch = f"{code} / {name}".strip(' /')
    md.currency = grab(r'Döviz Cinsi\s*:\s*(\S+)') or ''
    period = (
        grab(r'Dönemi\s*:\s*([\d./]+\s*-\s*[\d./]+)', source=text) or
        grab(r'Dönem\s*\(Tarih Aralığı\)\s*:\s*([\d./]+\s*-\s*[\d./]+)', source=text) or
        grab(r'Dönem[^:]*:\s*([\d./]+\s*-\s*[\d./]+)', source=text)
    )
    if period and '-' in period:
        a, b = period.split('-', 1)
        md.period_start, md.period_end = a.strip(), b.strip()
    bal = grab(r'Hesap Bakiyesi\s*:\s*([\-\d.,]+)', source=text)
    if bal:
        try:
            md.stated_balance = parse_tr_decimal(bal)
        except InvalidOperation:
            pass
    return md


def _normalize_date(date_str: str) -> str:
    return date_str.replace('.', '-')


def _hb_parse_transactions(full_text: str) -> tuple[list[TransactionRow], list[str]]:
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    current: Optional[TransactionRow] = None

    for raw in full_text.split('\n'):
        line = raw.rstrip('\r').rstrip()
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith('İşlem Tarihi') or stripped.startswith('Sayfa No'):
            continue
        if stripped in _HB_SKIP_EXACT:
            continue
        m = _HB_ROW_V1.match(line)
        if m:
            if current is not None:
                rows.append(current)
            date_str, amount_str, balance_str, desc = m.groups()
            try:
                amount = parse_tr_decimal(amount_str)
                balance = parse_tr_decimal(balance_str)
            except InvalidOperation as exc:
                warnings.append(f"Skipped malformed row at {date_str}: {exc}")
                current = None
                continue
            current = TransactionRow(date=date_str, amount=amount, balance=balance,
                                     description=desc.strip())
            continue
        m2 = _HB_ROW_V2.match(line)
        if m2:
            if current is not None:
                rows.append(current)
            date_str, amount_str, amount_sign, balance_str, balance_sign, desc = m2.groups()
            try:
                amount = parse_tr_decimal(amount_str)
                balance = parse_tr_decimal(balance_str)
            except InvalidOperation as exc:
                warnings.append(f"Skipped malformed row at {date_str}: {exc}")
                current = None
                continue
            if amount_sign == '-':
                amount = -amount
            if balance_sign == '-':
                balance = -balance
            current = TransactionRow(date=_normalize_date(date_str), amount=amount,
                                     balance=balance, description=desc.strip())
            continue
        if current is None:
            continue
        if stripped.startswith(_HB_SKIP_CONT_PREFIXES):
            continue
        if stripped.startswith('Ekstrenize') or stripped.startswith('Türkiye Halk Bankası'):
            continue
        current.description += stripped
    if current is not None:
        rows.append(current)
    for r in rows:
        r.receipt = _hb_extract_dekont(r.description)
    return rows, warnings


def parse_halkbank(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    metadata = _hb_extract_metadata(pdf_bytes)
    rows, warnings = _hb_parse_transactions(full_text)
    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 6. Akbank parser  (unchanged from original)
# =============================================================================

_AK_NUM = r'-?[\d.]+,\d{2}'
_AK_ROW = re.compile(
    rf'^(\d{{2}}\.\d{{2}}\.\d{{4}})\s+'
    rf'(\d{{2}}:\d{{2}})\s+'
    rf'\d{{2}}\.\d{{2}}\.\d{{4}}\s+'
    rf'(\d+)\s+'
    rf'({_AK_NUM})\s+'
    rf'({_AK_NUM})\s+'
    rf'([BA])\s+'
    rf'(.*)$'
)
_AK_SKIP_LINE_STARTS = (
    'HESAP HAREKET', 'Tarih :', 'Düzenleyen', 'Hesap Şube', 'Hesap No',
    'Döviz', 'TARİH', 'AKBANK',
)


def _ak_extract_metadata(pdf_bytes: bytes) -> StatementMetadata:
    md = StatementMetadata(bank='AKBANK')
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
    except Exception:
        return md

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, text)
        return m.group(1).strip() if m else None

    md.customer_name = grab(r'Ad Soyad\s*:\s*([^\n]+)') or ''
    md.account_no = grab(r'Hesap No\s*:\s*(\S+)') or ''
    md.iban = grab(r'IBAN\s*:\s*(\S+)') or ''
    branch_m = re.search(r'Hesap Şube\s*:\s*(.+?)(?:\s{2,}|\s+(?:IBAN|Ad Soyad|Tarih))', text)
    md.branch = branch_m.group(1).strip() if branch_m else (grab(r'Hesap Şube\s*:\s*([^\n]+)') or '')
    curr = grab(r'Döviz\s*:\s*([^\n]+)')
    if curr:
        parts = curr.split('-')
        md.currency = parts[-1].strip() if len(parts) > 1 else curr.strip()
    period = grab(r'Tarih Aralığı\s*:\s*([\d./]+\s*-\s*[\d./]+)')
    if period and '-' in period:
        a, b = period.split('-', 1)
        md.period_start, md.period_end = a.strip(), b.strip()
    return md


def _ak_parse_transactions(full_text: str) -> tuple[list[TransactionRow], list[str]]:
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    current: Optional[TransactionRow] = None
    for raw in full_text.split('\n'):
        line = raw.rstrip('\r').rstrip()
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith(_AK_SKIP_LINE_STARTS):
            continue
        if re.match(r'^\d+\s*/\s*\d+$', stripped):
            continue
        m = _AK_ROW.match(line)
        if m:
            if current is not None:
                rows.append(current)
            date_str, time_str, fis_no, amount_str, balance_str, ba, desc = m.groups()
            try:
                amount = parse_tr_decimal(amount_str)
                balance = parse_tr_decimal(balance_str)
            except InvalidOperation as exc:
                warnings.append(f"Skipped malformed row at {date_str}: {exc}")
                current = None
                continue
            amount = -abs(amount) if ba == 'B' else abs(amount)
            current = TransactionRow(
                date=_normalize_date(date_str), time=time_str,
                amount=amount, balance=balance,
                description=desc.strip(), receipt=fis_no,
            )
            continue
        if current is None:
            continue
        current.description += ' ' + stripped
    if current is not None:
        rows.append(current)
    rows.reverse()
    return rows, warnings


def parse_akbank(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    metadata = _ak_extract_metadata(pdf_bytes)
    rows, warnings = _ak_parse_transactions(full_text)
    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 7. Ziraat Bankası parser — supports BOTH vector PDFs and scanned/OCR PDFs
# =============================================================================

_ZR_DATE_RE = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')

# Strict Turkish-decimal money pattern: optional sign, optional thousands groups,
# mandatory comma followed by exactly 2 digits at end.
_TR_MONEY_RE = re.compile(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}$|^-?\d+,\d{2}$')

# Header / footer / metadata line prefixes to skip in OCR text
_ZR_OCR_SKIP_PREFIXES = (
    'Tarih', 'Fiş', 'Sayın', 'Müşteri', 'Müşter', 'Adres', 'HENDESE',
    'Şube', 'IBAN', 'Dönem', 'Day', 'Döviz', 'Borç', 'Alacak',
    'Taraflar', 'Merkez', 'Ticaret Sicil', 'www.', 'ğ', 'EEE',
)


def _normalize_ocr_number(s: str) -> str:
    """Fix common OCR mistakes in Turkish-formatted numbers.

    Turkish money: `1.234.567,89` (period = thousands separator, comma = decimal).
    OCR often reads the decimal comma as a period, producing e.g. `75.001.60`
    or `37.51`. Rule: if the number has no comma but ends in `.XX`, that final
    period IS the decimal — convert it to a comma.
    """
    s = s.strip()
    if not s or ',' in s:
        return s
    if re.search(r'\.\d{2}$', s):
        idx = s.rfind('.')
        s = s[:idx] + ',' + s[idx + 1:]
    return s


def _try_repair_no_separator(tok: str) -> Optional[str]:
    """Repair tokens where OCR dropped the decimal entirely.

    e.g. `3751` for `37,51`. Only attempted on pure digit strings of length ≥ 3.
    The repaired number is later validated by the balance chain — if it makes
    the chain consistent we keep it (and surface a warning); if not, the user
    sees a clear flag instead of a silent miscount.
    """
    s = tok.strip()
    sign = ''
    if s.startswith('-'):
        sign = '-'
        s = s[1:]
    if not s.isdigit() or len(s) < 3 or len(s) > 12:
        return None
    return f"{sign}{s[:-2]},{s[-2:]}"


def _is_money(tok: str) -> bool:
    return bool(_TR_MONEY_RE.match(tok))


def _ziraat_extract_metadata(full_text: str) -> StatementMetadata:
    md = StatementMetadata(bank='ZİRAAT')

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, full_text, re.IGNORECASE)
        return m.group(1).strip() if m else None

    cust_raw = grab(r'Say[ıi]n\s*:?\s*([^\n]+)') or ''
    md.customer_name = re.split(r'\s{3,}|\s+[Şş]ube', cust_raw)[0].strip()
    # Strip any leading "O:" or "0:" left over from OCR'd "Sayın :" prefix
    md.customer_name = re.sub(r'^[O0]\s*:\s*', '', md.customer_name).strip()
    md.account_no = grab(r'M[üu][şs]ter[i/]+Hesap No\s*:\s*(\S+)') or grab(r'Hesap No\s*:\s*(\S+)') or ''
    md.iban = grab(r'IBAN\s*:\s*(\S+)') or ''
    md.branch = grab(r'[Şş]ube Kodu\s*:\s*([^\n\t]+)') or ''
    md.currency = grab(r'D[öo]viz Cinsi\s*:\s*(\S+)') or ''
    # Period: OCR often misreads ':' as '1', mangles digits, and turns inner
    # date periods into spaces. Find the first "two-dates-with-hyphen"
    # pattern within ~30 chars after 'Dönem'.
    pm = re.search(
        r'D[öo]nem.{0,30}?(\d{2}[.\-\s]\d{2}[.\-\s]\d{4})\s*[-–]\s*(\d{2}[.\-\s]\d{2}[.\-\s]\d{4})',
        full_text,
        re.DOTALL,
    )
    if pm:
        # Validate & repair OCR digit confusions (e.g. day '50' -> '30').
        md.period_start = _fix_ocr_date(pm.group(1).strip()) or ''
        md.period_end   = _fix_ocr_date(pm.group(2).strip()) or ''
    return md


def _ziraat_parse_vector(pdf_bytes: bytes) -> tuple[list[TransactionRow], list[str]]:
    """Parse Ziraat statement when PDF has an embedded text layer (uses tables)."""
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    for trow in table:
                        if not trow or len(trow) < 5:
                            continue
                        date_val = (trow[0] or '').strip()
                        fis_no = (trow[1] or '').strip()
                        desc = ' '.join((trow[2] or '').split())
                        tutar_s = (trow[3] or '').strip()
                        bakiye_s = (trow[4] or '').strip()
                        if not _ZR_DATE_RE.match(date_val):
                            continue
                        if not tutar_s or not bakiye_s:
                            continue
                        try:
                            amount = parse_tr_decimal(tutar_s)
                            balance = parse_tr_decimal(bakiye_s)
                        except (InvalidOperation, Exception) as exc:
                            warnings.append(f"Skipped Ziraat row at {date_val}: {exc}")
                            continue
                        rows.append(TransactionRow(
                            date=_normalize_date(date_val),
                            amount=amount, balance=balance,
                            description=desc, receipt=fis_no,
                        ))
    except Exception as exc:
        warnings.append(f"Ziraat table extraction error: {exc}")
    rows.reverse()  # newest-first → chronological
    return rows, warnings


def _ziraat_normalize_fis(fis: str) -> str:
    """Clean OCR artifacts from Ziraat Fiş No.

    Legitimate Ziraat fiş numbers are 6 chars: 'F' + 5 digits (e.g. F12750)
    or alphanumeric (e.g. FBWW24). OCR commonly produces:
      - leading junk: '(OF12750', 'oOF12679', '(F10836'
      - I/1 and O/0 confusion: 'FI7539' (real F17539), 'FO9116' (real F09116)
      - spurious extra char: 'FI12679' (real F12679, I is junk insert)

    Strategy: strip leading non-F junk; if result starts F[IlO]:
      - len 6: convert position-1 char (I/l->1, O->0) -- preserves length
      - len 7: drop the spurious char -- restores canonical 6-char length
    """
    if not fis:
        return fis
    fis = re.sub(r'^[(\s]+', '', fis)
    fis = re.sub(r'^[oO0]+(?=F)', '', fis)  # strip leading O/0/o ONLY if F follows
    fis = re.sub(r'[)\s]+$', '', fis)
    m = re.match(r'^(F)([IlO])(\d{4,5})$', fis)
    if m:
        body = m.group(3)
        if len(fis) == 6:
            fixed = {'I': '1', 'l': '1', 'O': '0'}[m.group(2)]
            fis = m.group(1) + fixed + body
        elif len(fis) == 7:
            fis = m.group(1) + body
    return fis


def _ziraat_parse_ocr(pdf_bytes: bytes) -> tuple[list[TransactionRow], list[str], str]:
    """Parse Ziraat statement when PDF is image-only (uses OCR with positional data).

    Returns (rows, warnings, full_ocr_text). The OCR text is also used to extract
    metadata (customer name, IBAN, period, etc.).
    """
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    full_text_parts: list[str] = []

    pages = _ocr_pages(pdf_bytes, dpi=400, lang='tur')
    for page_data in pages:
        full_text_parts.append(page_data['text'])

        pending_desc: list[str] = []
        for line in page_data['lines']:
            tokens = line['tokens']
            if not tokens:
                continue

            if _ZR_DATE_RE.match(tokens[0]):
                date = tokens[0]
                # Find money tokens AFTER normalization & repair
                normalized: list[tuple[int, str, bool]] = []  # (idx, value, was_repaired)
                repair_log: list[tuple[str, str]] = []
                for i, t in enumerate(tokens):
                    if i == 0:  # date itself — never treat as money
                        normalized.append((i, t, False))
                        continue
                    n = _normalize_ocr_number(t)
                    if _is_money(n):
                        normalized.append((i, n, n != t))
                    else:
                        repaired = _try_repair_no_separator(t)
                        if repaired and _is_money(repaired):
                            normalized.append((i, repaired, True))
                            repair_log.append((t, repaired))
                        else:
                            normalized.append((i, t, False))

                money_idxs = [(i, v, r) for i, v, r in normalized
                              if i > 0 and _is_money(v)]
                if len(money_idxs) >= 2:
                    tutar_idx, tutar_s, tutar_repaired = money_idxs[-2]
                    _bal_idx, bakiye_s, bakiye_repaired = money_idxs[-1]
                    # Fiş No: first token after the date that doesn't look like OCR
                    # junk. OCR often inserts noise tokens like "(o" / "(O" / "o"
                    # before the real fiş — skip them.
                    fis = ''
                    junk_re = re.compile(r'^[(\s]*[oO0]?[)\s]*$')
                    for ti in range(1, tutar_idx):
                        cand = re.sub(r'^[(\s]*[oO0]?\s*', '', tokens[ti]).strip()
                        cand = re.sub(r'\s*[)\s]*$', '', cand).strip()
                        if not cand or junk_re.match(cand):
                            continue
                        if len(cand) >= 3:
                            fis = _ziraat_normalize_fis(cand)
                            break
                    desc_inline = ' '.join(
                        tokens[ti + 1:tutar_idx]
                    ).strip() if fis else ' '.join(tokens[1:tutar_idx]).strip()
                    full_desc_parts = pending_desc + (
                        [desc_inline] if desc_inline else []
                    )
                    full_desc = ' '.join(full_desc_parts).strip()
                    pending_desc = []
                    try:
                        amount = parse_tr_decimal(tutar_s)
                        balance = parse_tr_decimal(bakiye_s)
                    except (InvalidOperation, Exception) as exc:
                        warnings.append(f"Skipped Ziraat OCR row at {date}: {exc}")
                        continue
                    rows.append(TransactionRow(
                        date=_normalize_date(date),
                        amount=amount, balance=balance,
                        description=full_desc, receipt=fis,
                    ))
                    if tutar_repaired or bakiye_repaired:
                        for orig, fixed in repair_log:
                            if fixed in (tutar_s, bakiye_s):
                                warnings.append(
                                    f"OCR repair on {date}: '{orig}' → '{fixed}'. "
                                    f"Verify against the source PDF."
                                )
                else:
                    warnings.append(
                        f"Ziraat OCR row at {date}: only {len(money_idxs)} valid "
                        f"number(s), expected 2. Tokens: {tokens}. ROW SKIPPED."
                    )
                    pending_desc = []
            else:
                # Not a transaction line — could be description continuation OR
                # header/footer/boilerplate. Skip known boilerplate; otherwise
                # buffer as a possible description continuation.
                line_text = ' '.join(tokens)
                if any(line_text.startswith(p) for p in _ZR_OCR_SKIP_PREFIXES):
                    continue
                pending_desc.append(line_text)

    rows.reverse()  # PDF lists newest-first → chronological
    return rows, warnings, '\n'.join(full_text_parts)


def parse_ziraat(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    """Ziraat parser. Picks vector or OCR path based on whether the PDF
    has an embedded text layer."""
    is_scanned = _is_pdf_scanned(pdf_bytes)
    if is_scanned:
        if not OCR_AVAILABLE:
            md = StatementMetadata(bank='ZİRAAT')
            return StatementResult(
                source_filename=source_filename, metadata=md, rows=[],
                parser_warnings=[
                    "This Ziraat PDF is image-only (scanned). OCR is required "
                    "but pytesseract / pdf2image / tesseract are not installed. "
                    f"Details: {OCR_IMPORT_ERROR or 'see deployment notes at top of file.'}"
                ],
            )
        rows, raw_msgs, ocr_text = _ziraat_parse_ocr(pdf_bytes)
        metadata = _ziraat_extract_metadata(ocr_text)
        # Split: OCR-detected note + repair notes are informational; everything
        # else (skipped rows, parse errors) is actionable.
        info: list[str] = [
            "Scanned/image-only PDF detected — used OCR (Turkish). "
            "Verify amounts against the source PDF."
        ]
        warnings: list[str] = []
        for m in raw_msgs:
            (info if m.startswith('OCR repair') else warnings).append(m)
        return StatementResult(
            source_filename=source_filename, metadata=metadata,
            rows=rows, parser_warnings=warnings, parser_info=info,
        )
    rows, warnings = _ziraat_parse_vector(pdf_bytes)
    metadata = _ziraat_extract_metadata(full_text)
    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 7B. Yapı Kredi parser
# =============================================================================
#
# Format characteristics:
#   * Date column: DD/MM/YYYY (slashes)
#   * Optional time column: HH:MM:SS
#   * Amounts have ` TL` suffix; sign always glued to the digits (`-18.320,95 TL`)
#   * Description sometimes wraps to lines PRECEDING the date row (when the
#     Açıklama column has 2-3 lines of text, pdfplumber emits them just above
#     the row that holds the date).
#   * Statement is newest-first; we reverse to chronological at the end.

_YK_DATE_RE = re.compile(r'^(\d{2}/\d{2}/\d{4})\b')
_YK_ROW_RE = re.compile(
    r'^(\d{2}/\d{2}/\d{4})\s+'                       # date
    r'(?:(\d{2}:\d{2}:\d{2})\s+)?'                   # optional time
    r'(.+?)\s+'                                       # description (reluctant)
    r'(-?[\d.]+,\d{2})\s*TL\s+'                      # amount + TL
    r'(-?[\d.]+,\d{2})\s*TL\s*$'                     # balance + TL
)

_YK_SKIP_PREFIXES = (
    'Hesap Hareketleri', 'Müşteri Adı', 'Müşteri Numarası', 'Şube',
    'Hesap Adı', 'IBAN', 'Kullanılabilir Bakiye', 'Tarih Aralığı',
    'Tarih Saat', 'Tarih ', 'Yapı ve Kredi', 'www.yapikredi',
    'Ticaret Sicil', 'Mersis No', 'İşletmenin Merkezi', 'Blok 34330',
    'T: ', 'F: ',
)


def _yk_extract_metadata(text: str) -> StatementMetadata:
    md = StatementMetadata(bank='YAPIKREDI')

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, text[:2000], re.IGNORECASE)
        return m.group(1).strip() if m else None

    md.customer_name = grab(r'Müşteri Adı(?:\s*Soyadı)?\s*:\s*([^\n]+)') or ''
    md.customer_no = grab(r'Müşteri Numarası\s*:\s*([^\n]+)') or ''
    md.branch = grab(r'Şube\s*:\s*([^\n]+)') or ''
    md.iban = grab(r'IBAN(?:/Hesap No)?\s*:\s*(TR\S+)') or ''
    md.currency = 'TL'  # YK shows "Vadesiz TL Hesabı" — always TRY in our test files
    period = grab(r'Tarih Aralığı\s*:\s*(\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4})')
    if period and '-' in period:
        a, b = period.split('-', 1)
        md.period_start, md.period_end = a.strip(), b.strip()
    return md


def _yk_norm_date(s: str) -> str:
    """13/05/2026 -> 13-05-2026."""
    return s.replace('/', '-')


def parse_yapikredi(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    metadata = _yk_extract_metadata(full_text)
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    pre_desc: list[str] = []

    for raw in full_text.split('\n'):
        line = raw.rstrip('\r').rstrip()
        stripped = line.strip()
        if not stripped:
            pre_desc = []
            continue
        # Skip headers/footers
        if any(stripped.startswith(p) for p in _YK_SKIP_PREFIXES):
            pre_desc = []
            continue

        m = _YK_ROW_RE.match(stripped)
        if m:
            date_s, time_s, desc, amt_s, bal_s = m.groups()
            try:
                amount = parse_tr_decimal(amt_s)
                balance = parse_tr_decimal(bal_s)
            except (InvalidOperation, Exception) as exc:
                warnings.append(f"Skipped Yapı Kredi row at {date_s}: {exc}")
                pre_desc = []
                continue
            full_desc = ' '.join(pre_desc + [desc]).strip()
            # Trim any trailing dashes that are description punctuation
            # (e.g. 'GELEN FAST - Hakan Gecü -') so the desc reads cleanly.
            full_desc = re.sub(r'\s+-\s*$', '', full_desc)
            rows.append(TransactionRow(
                date=_yk_norm_date(date_s), time=time_s or '',
                amount=amount, balance=balance, description=full_desc,
            ))
            pre_desc = []
        elif _YK_DATE_RE.match(stripped):
            # Looks like a date row but didn't match — incomplete extraction.
            warnings.append(f"Yapı Kredi: date row didn't match expected layout: {stripped[:120]!r}")
            pre_desc = []
        else:
            # Continuation line (description spilled above the row)
            pre_desc.append(stripped)

    rows.reverse()  # statement is newest-first
    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 7C. Kuveyt Türk parser
# =============================================================================
#
# Format characteristics:
#   * Date column: DD.MM.YYYY (dots)
#   * Reference code: 5-char uppercase alphanumeric (A6DJ0, A01PG, ASHSZ, ...)
#   * Words in description often glued together: 'Gönderen:HakanGecü,Alıcı:GECO...'
#     (PDF extraction artifact; we accept it as-is).
#   * Tutar/Bakiye use Turkish comma decimals (e.g. 41.712,07).
#   * NEGATIVE-SIGN QUIRK: when the negative amount is right-aligned in its
#     column and wraps, the '-' character lands on the line ABOVE the date
#     row, and the numeric value lands on the line BELOW it. So a single
#     logical row may span 3 text lines:
#         '-'
#         '19.01.2026 A01PG <desc> 0,00'
#         '41.712,07'
#     where the actual amount is -41.712,07 and the balance on the date line
#     (0,00) is the second column. We detect this by counting trailing
#     numbers on the date row.
#   * Statement is oldest-first → no reverse.

_KT_NUM = r'-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}'
_KT_DATE_REF_RE = re.compile(
    rf'^(\d{{2}}\.\d{{2}}\.\d{{4}})\s+([A-Z0-9]{{4,8}})\s+(.+)$'
)
# Description may be empty (entire desc was on lines ABOVE this date row).
_KT_TWO_NUMS_TAIL = re.compile(rf'^(.*?)\s*({_KT_NUM})\s+({_KT_NUM})\s*$')
_KT_ONE_NUM_TAIL  = re.compile(rf'^(.*?)\s*({_KT_NUM})\s*$')
_KT_BARE_NUM_RE = re.compile(rf'^({_KT_NUM})\s*$')
_KT_BARE_DASH_RE = re.compile(r'^-\s*$')

_KT_SKIP_PREFIXES = (
    'Hesap Bilgileri', 'HesapBilgileri', 'IBAN', 'Döviz', 'DövizTürü',
    'Dönem', 'Hesap', 'Toplam', 'Sayın', 'HESAP ÖZETİ', 'HESAPÖZETİ',
    'Oluşturulduğu', 'KUVEYT', 'Büyükdere', 'Ticaret', 'Mersis',
    'İşlem', 'Açıklama', 'Devir', 'Tarihi', 'Referans',
    'www.kuveytturk', 'Bakiyesi',
)


def _kt_extract_metadata(text: str) -> StatementMetadata:
    md = StatementMetadata(bank='KUVEYTTURK')

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, text[:2500], re.IGNORECASE)
        return m.group(1).strip() if m else None

    md.customer_name = grab(r'Sayın\s+([A-ZÇĞİÖŞÜ][^\n]+?)(?:\s*$|\n)') or ''
    md.account_no = grab(r'Hesap\s*Bilgileri\s*:\s*(\S+)') or ''
    # IBAN may contain spaces every 4 chars — collapse
    iban = grab(r'IBAN\s*:\s*(TR[\d\s]+)') or ''
    md.iban = re.sub(r'\s+', '', iban) if iban else ''
    period = grab(r'Dönem\s*:\s*(\d{2}\.\d{2}\.\d{4}\s*[\u2013\-]\s*\d{2}\.\d{2}\.\d{4})')
    if period:
        parts = re.split(r'\s*[\u2013\-]\s*', period, maxsplit=1)
        if len(parts) == 2:
            md.period_start, md.period_end = parts[0].strip(), parts[1].strip()
    md.currency = 'TL'
    return md


def parse_kuveytturk(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    metadata = _kt_extract_metadata(full_text)
    rows: list[TransactionRow] = []
    warnings: list[str] = []

    raw_lines = [l.strip() for l in full_text.split('\n')]
    lines: list[str] = [l for l in raw_lines if l]  # drop empties for sequential lookup

    i = 0
    pre_desc: list[str] = []
    pending_negative = False
    while i < len(lines):
        line = lines[i]

        if _KT_BARE_DASH_RE.match(line):
            pending_negative = True
            i += 1
            continue

        if any(line.startswith(p) for p in _KT_SKIP_PREFIXES):
            pre_desc = []
            pending_negative = False
            i += 1
            continue

        m = _KT_DATE_REF_RE.match(line)
        if not m:
            # Continuation description line BEFORE the date row
            pre_desc.append(line)
            i += 1
            continue

        date_s, ref, rest = m.groups()
        amount: Optional[Decimal] = None
        balance: Optional[Decimal] = None
        desc_inline = ''

        # Case A: date line ends with two numbers → amount + balance in-line.
        two = _KT_TWO_NUMS_TAIL.match(rest)
        if two:
            desc_inline = two.group(1)
            try:
                amount = parse_tr_decimal(two.group(2))
                balance = parse_tr_decimal(two.group(3))
            except (InvalidOperation, Exception) as exc:
                warnings.append(f"Kuveyt Türk: bad numbers at {date_s}: {exc}")
                pre_desc = []; pending_negative = False; i += 1
                continue
            # If we already saw a standalone '-' on the previous line, the amount
            # in this 2-number form is sometimes still meant to be negative —
            # but only when the amount itself has no glued sign. Keep simple:
            # trust the sign embedded in the number.
        else:
            # Case B: date line ends with ONE number → that's the BALANCE.
            # The AMOUNT is on the next line (with sign on a line before date).
            one = _KT_ONE_NUM_TAIL.match(rest)
            if one and i + 1 < len(lines) and _KT_BARE_NUM_RE.match(lines[i + 1]):
                desc_inline = one.group(1)
                try:
                    balance = parse_tr_decimal(one.group(2))
                    amount = parse_tr_decimal(lines[i + 1])
                except (InvalidOperation, Exception) as exc:
                    warnings.append(f"Kuveyt Türk: bad numbers at {date_s}: {exc}")
                    pre_desc = []; pending_negative = False; i += 2
                    continue
                if pending_negative:
                    amount = -amount
                i += 1  # consume the amount line
            else:
                warnings.append(
                    f"Kuveyt Türk: couldn't parse row at {date_s}: {line[:120]!r}"
                )
                pre_desc = []; pending_negative = False; i += 1
                continue

        full_desc = ' '.join(pre_desc + [desc_inline]).strip()
        # Strip trailing standalone '-' (the visual negative sign that got
        # captured into the description column when sign was on a same line).
        full_desc = re.sub(r'\s+-\s*$', '', full_desc)

        rows.append(TransactionRow(
            date=_normalize_date(date_s),
            amount=amount, balance=balance,
            description=full_desc, receipt=ref,
        ))
        pre_desc = []
        pending_negative = False
        i += 1

    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 7D. Türkiye Finans Katılım parser
# =============================================================================
#
# Format characteristics:
#   * Date: 'D.MM.YYYY' — single-digit day allowed (no leading zero!)
#   * Reference: 5-char uppercase alphanumeric (XAQ5D, XK2PX, ...)
#   * Amounts in ENGLISH format: period decimal, 4 decimal places, NO thousands
#     separator: '-11171.9600 0.0000'.
#   * Multi-column wrap: the 'İşlem' column ('Diğer Bankacılık İşlemleri') and
#     the 'Açıklama' column each wrap to their own continuation lines. The
#     first line of a transaction contains date+ref+partial-fields+amount+balance.
#     Continuation lines below contain the rest of those fields, interleaved.
#     We accept the messy description and validate via balance chain.
#   * Statement is newest-first → reverse to chronological.

_TF_NUM = r'-?\d+\.\d{2,4}'
_TF_FIRST_LINE = re.compile(
    rf'^(\d{{1,2}}\.\d{{1,2}}\.\d{{4}})\s+'    # date
    rf'(\S+)\s+'                                # ref
    rf'(.+?)\s+'                                 # middle text (reluctant)
    rf'({_TF_NUM})\s+'                          # amount
    rf'({_TF_NUM})\s*$'                          # balance
)
_TF_DATE_PREFIX = re.compile(r'^\d{1,2}\.\d{1,2}\.\d{4}\s')

_TF_SKIP_PREFIXES = (
    'Hesap Hareketleri', 'Müşteri', 'Döviz', 'Şube', 'Hesap Adı',
    'IBAN', 'Kullanılabilir Bakiye', 'Tarih Aralığı',
    'Tarih İşlem', 'Tarih ',
)


def _tf_extract_metadata(text: str) -> StatementMetadata:
    md = StatementMetadata(bank='TURKIYEFINANS')

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, text[:2000], re.IGNORECASE)
        return m.group(1).strip() if m else None

    md.customer_name = grab(r'Müşteri Adı Soyadı\s+([^\n]+?)(?:\s*$|\n)') or ''
    md.customer_no = grab(r'Müşteri Numarası\s+([^\n]+?)(?:\s*$|\n)') or ''
    md.branch = grab(r'Şube\s+([^\n]+?)(?:\s*$|\n)') or ''
    md.iban = grab(r'IBAN(?:/Hesap No)?\s+(TR\S+)') or ''
    md.currency = grab(r'Döviz Cinsi\s+(\S+)') or 'TL'
    period = grab(r'Tarih Aralığı\s+(\d{2}\.\d{2}\.\d{4}\s*-\s*\d{2}\.\d{2}\.\d{4})')
    if period and '-' in period:
        a, b = period.split('-', 1)
        md.period_start, md.period_end = a.strip(), b.strip()
    return md


def _tf_norm_date(s: str) -> str:
    """1.04.2026 or 01.04.2026 -> 01-04-2026."""
    parts = s.split('.')
    if len(parts) == 3:
        try:
            return f'{int(parts[0]):02d}-{int(parts[1]):02d}-{parts[2]}'
        except ValueError:
            pass
    return s


def parse_turkiyefinans(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    metadata = _tf_extract_metadata(full_text)
    rows: list[TransactionRow] = []
    warnings: list[str] = []

    lines = full_text.split('\n')
    # First pass: find rows that match the date+ref+amount+balance pattern.
    # Continuation lines are appended to the previous row's description until
    # the next date-prefixed line is seen.
    current_row: Optional[TransactionRow] = None
    pending_desc_extra: list[str] = []

    def _flush_current() -> None:
        nonlocal current_row, pending_desc_extra
        if current_row is not None:
            if pending_desc_extra:
                extras = ' '.join(s for s in pending_desc_extra if s)
                current_row.description = (
                    f'{current_row.description} {extras}'.strip()
                )
            rows.append(current_row)
        current_row = None
        pending_desc_extra = []

    for raw in lines:
        line = raw.rstrip('\r').rstrip()
        stripped = line.strip()
        if not stripped:
            continue
        if any(stripped.startswith(p) for p in _TF_SKIP_PREFIXES):
            _flush_current()
            continue

        m = _TF_FIRST_LINE.match(stripped)
        if m:
            _flush_current()
            date_s, ref, desc_part, amt_s, bal_s = m.groups()
            try:
                amount = Decimal(amt_s)
                balance = Decimal(bal_s)
            except (InvalidOperation, Exception) as exc:
                warnings.append(f"Türkiye Finans: bad numbers at {date_s}: {exc}")
                continue
            current_row = TransactionRow(
                date=_tf_norm_date(date_s),
                amount=amount, balance=balance,
                description=desc_part.strip(), receipt=ref,
            )
            continue

        # Not a transaction-start line.
        if _TF_DATE_PREFIX.match(stripped):
            # Starts with date but the full pattern didn't match — log & skip.
            warnings.append(
                f"Türkiye Finans: date line failed pattern: {stripped[:120]!r}"
            )
            _flush_current()
            continue

        # Continuation of current row's description (or noise we'll discard
        # if there's no current row).
        if current_row is not None:
            pending_desc_extra.append(stripped)

    _flush_current()
    rows.reverse()  # newest-first → chronological

    # Clean up descriptions: drop the constant "Diğer Bankacılık İşlemleri"
    # category prefix tokens that get interleaved with the description.
    for r in rows:
        d = r.description
        # Remove "Diğer", "Bankacılık", "İşlemleri" boilerplate tokens with
        # only whitespace between them, anywhere in the string.
        d = re.sub(r'\bDiğer\s+Bankacılık\s+İşlemleri\b', '', d)
        # Also strip these tokens individually if they're standalone (the
        # PDF interleaves them column-wise so we sometimes see them alone).
        d = re.sub(r'\b(Diğer|Bankacılık|İşlemleri)\b', '', d)
        # Collapse whitespace.
        d = re.sub(r'\s+', ' ', d).strip()
        r.description = d

    return StatementResult(source_filename=source_filename, metadata=metadata,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 8. Generic fallback parser  (unchanged from original)
# =============================================================================

_GEN_DATE = r'\d{2}[.\-/]\d{2}[.\-/]\d{4}'
_GEN_TIME = r'\d{2}:\d{2}'
_GEN_NUM = r'-?[\d.]+,\d{2}'

_GEN_PATTERNS = [
    ('akbank_like', re.compile(
        rf'^({_GEN_DATE})\s+({_GEN_TIME})\s+{_GEN_DATE}\s+\d+\s+'
        rf'({_GEN_NUM})\s+({_GEN_NUM})\s+([BA])\s+(.*)$'
    )),
    ('halkbank_v2_like', re.compile(
        rf'^({_GEN_DATE})\s+{_GEN_DATE}\s+'
        rf'([\d.]+,\d{{2}})\s+([+-])\s+'
        rf'([\d.]+,\d{{2}})\s+([+-])\s+(.*)$'
    )),
    ('halkbank_v1_like', re.compile(
        rf'^({_GEN_DATE})\s+({_GEN_NUM})\s+({_GEN_NUM})\s+(.*)$'
    )),
]
_GEN_SKIP_STARTS = (
    'HESAP', 'TARİH', 'TARIH', 'İŞLEM', 'SAYFA', 'AKBANK', 'HALKBANK',
    'GARANTİ', 'ZİRAAT', 'MÜŞTERİ', 'MUSTERI', 'IBAN', 'DÖVIZ', 'DOVIZ',
    'DÜZENLEYEN', 'ŞUBE', 'SUBE', 'DÖNEM', 'DONEM', 'TCKN', 'Müşteri',
    'Hesap', 'Şube', 'Döviz', 'Dönem', 'Bakiye', 'Bloke', 'Toplam',
    'Türkiye', 'Ekstrenize', 'yerine',
)


def _generic_parse_transactions(full_text: str) -> tuple[list[TransactionRow], list[str]]:
    rows: list[TransactionRow] = []
    warnings: list[str] = []
    current: Optional[TransactionRow] = None
    for raw in full_text.split('\n'):
        line = raw.rstrip('\r').rstrip()
        stripped = line.strip()
        if not stripped:
            continue
        matched = False
        for pat_name, pat_re in _GEN_PATTERNS:
            m = pat_re.match(line)
            if not m:
                continue
            if current is not None:
                rows.append(current)
            groups = m.groups()
            if pat_name == 'akbank_like':
                date_str, time_str, amount_str, balance_str, ba, desc = groups
                try:
                    amount = parse_tr_decimal(amount_str)
                    balance = parse_tr_decimal(balance_str)
                except InvalidOperation:
                    current = None; matched = True; break
                amount = -abs(amount) if ba == 'B' else abs(amount)
                current = TransactionRow(date=_normalize_date(date_str), time=time_str,
                                         amount=amount, balance=balance, description=desc.strip())
            elif pat_name == 'halkbank_v2_like':
                date_str, amount_str, amount_sign, balance_str, balance_sign, desc = groups
                try:
                    amount = parse_tr_decimal(amount_str)
                    balance = parse_tr_decimal(balance_str)
                except InvalidOperation:
                    current = None; matched = True; break
                if amount_sign == '-': amount = -amount
                if balance_sign == '-': balance = -balance
                current = TransactionRow(date=_normalize_date(date_str), amount=amount,
                                         balance=balance, description=desc.strip())
            elif pat_name == 'halkbank_v1_like':
                date_str, amount_str, balance_str, desc = groups
                try:
                    amount = parse_tr_decimal(amount_str)
                    balance = parse_tr_decimal(balance_str)
                except InvalidOperation:
                    current = None; matched = True; break
                current = TransactionRow(date=_normalize_date(date_str), amount=amount,
                                         balance=balance, description=desc.strip())
            matched = True
            break
        if not matched and current is not None:
            if any(stripped.upper().startswith(s.upper()) for s in _GEN_SKIP_STARTS):
                continue
            if re.match(r'^\d+\s*/\s*\d+$', stripped):
                continue
            current.description += ' ' + stripped
    if current is not None:
        rows.append(current)
    return rows, warnings


def _generic_block_parse(full_text: str) -> tuple[list[TransactionRow], list[str]]:
    """Block-accumulation fallback for unknown banks.

    Strategy: many bank statements have rows that wrap to multiple lines due
    to long descriptions. We scan for lines beginning with a date and end
    the current block when we hit either (a) the next date-prefixed line,
    or (b) end of text. From each block we try to extract trailing
    'amount + balance' (Turkish-decimal pairs) and treat the rest as
    description.

    Tries both Turkish (`1.234,56`) and English (`1234.56` / `1234.5600`)
    decimal formats, with optional `TL` / `TRY` suffix.

    Returns rows in their original order from the PDF — the caller decides
    whether to reverse.
    """
    # Date prefix variants — most TR statements use one of these.
    date_prefix_re = re.compile(
        r'^(\d{1,2}[./-]\d{1,2}[./-]\d{4})\b'
    )
    # Trailing amount + balance patterns, tried in order.
    tail_patterns = [
        # Turkish decimals with TL/TRY suffix
        re.compile(r'(-?[\d.]+,\d{2})\s*(?:TL|TRY)\s+(-?[\d.]+,\d{2})\s*(?:TL|TRY)\s*$'),
        # Turkish decimals plain
        re.compile(r'(-?[\d.]+,\d{2})\s+(-?[\d.]+,\d{2})\s*$'),
        # English decimals (2-4 dp)
        re.compile(r'(-?\d+(?:,\d{3})*\.\d{2,4})\s+(-?\d+(?:,\d{3})*\.\d{2,4})\s*$'),
    ]

    def _parse_number(s: str) -> Optional[Decimal]:
        s = s.strip()
        # Determine format: Turkish if it has a comma followed by 2 digits at end
        # and dots-as-thousands; English if it has a period followed by 2-4
        # digits at end.
        if re.search(r',\d{2}$', s):
            try:
                return parse_tr_decimal(s)
            except Exception:
                return None
        # English: strip commas, parse as float
        try:
            return Decimal(s.replace(',', ''))
        except (InvalidOperation, Exception):
            return None

    lines = full_text.split('\n')
    blocks: list[list[str]] = []
    current: list[str] = []
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        if date_prefix_re.match(line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)
    if current:
        blocks.append(current)

    rows: list[TransactionRow] = []
    warnings: list[str] = []
    for block in blocks:
        joined = ' '.join(block)
        # Find the first date in the block
        dm = date_prefix_re.match(joined)
        if not dm:
            continue
        date_s = dm.group(1)
        # Try each tail pattern
        amount = None
        balance = None
        desc = joined[dm.end():].strip()
        for pat in tail_patterns:
            tm = pat.search(joined)
            if tm:
                a = _parse_number(tm.group(1))
                b = _parse_number(tm.group(2))
                if a is not None and b is not None:
                    amount, balance = a, b
                    desc = joined[dm.end():tm.start()].strip()
                    break
        if amount is None or balance is None:
            continue
        # Normalize date separator to '-'
        date_n = re.sub(r'[./]', '-', date_s)
        # Zero-pad single-digit day/month
        parts = date_n.split('-')
        if len(parts) == 3:
            try:
                date_n = f'{int(parts[0]):02d}-{int(parts[1]):02d}-{parts[2]}'
            except ValueError:
                pass
        rows.append(TransactionRow(
            date=date_n, amount=amount, balance=balance,
            description=desc,
        ))
    return rows, warnings


def parse_generic(pdf_bytes: bytes, full_text: str, source_filename: str) -> StatementResult:
    md = StatementMetadata(bank='UNKNOWN')

    def grab(pattern: str) -> Optional[str]:
        m = re.search(pattern, full_text[:3000])
        return m.group(1).strip() if m else None

    md.customer_name = grab(r'(?:Ad Soyad|Müşteri Adı|Ünvanı?)\s*:\s*([^\n]+)') or ''
    md.account_no = grab(r'Hesap No\s*:\s*(\S+)') or ''
    md.iban = grab(r'IBAN\s*:\s*(\S+)') or ''
    period = grab(r'(?:Tarih Aralığı|Dönemi?|Dönem)\s*[:(]\s*([\d./]+\s*-\s*[\d./]+)')
    if period and '-' in period:
        a, b = period.split('-', 1)
        md.period_start, md.period_end = a.strip(), b.strip()

    # Strategy A: strict line-by-line patterns (the original generic logic).
    rows_a, warnings_a = _generic_parse_transactions(full_text)
    # Strategy B: block-accumulation (forgiving of wrapped rows).
    rows_b, warnings_b = _generic_block_parse(full_text)

    # Pick whichever extracted more rows and validates better. Prefer the one
    # with no balance issues; tiebreak on row count.
    cand_a_issues = len(validate_balance_chain(rows_a)) if rows_a else 999
    cand_b_issues = len(validate_balance_chain(rows_b)) if rows_b else 999

    def _score(rows: list, issues: int) -> tuple[int, int]:
        # Lower issues wins, then higher row count wins.
        return (issues, -len(rows))

    if _score(rows_b, cand_b_issues) < _score(rows_a, cand_a_issues):
        rows, warnings, strategy = rows_b, warnings_b, 'block-accumulation'
    else:
        rows, warnings, strategy = rows_a, warnings_a, 'line-by-line'

    warnings.insert(0,
        f"⚠ Used generic fallback parser ({strategy}) — bank not recognized. "
        f"Results may be incomplete. Please verify carefully."
    )

    if len(rows) >= 2:
        try:
            first_d = pd.to_datetime(rows[0].date, format='%d-%m-%Y', errors='coerce')
            last_d = pd.to_datetime(rows[-1].date, format='%d-%m-%Y', errors='coerce')
            if first_d is not None and last_d is not None and first_d > last_d:
                rows.reverse()
                warnings.append("Rows appeared reverse-chronological; reversed to oldest-first.")
        except Exception:
            pass
    return StatementResult(source_filename=source_filename, metadata=md,
                           rows=rows, parser_warnings=warnings)


# =============================================================================
# 9. Parser registry
# =============================================================================

PARSERS: dict[str, Callable[[bytes, str, str], StatementResult]] = {
    'HALKBANK':      parse_halkbank,
    'AKBANK':        parse_akbank,
    'ZIRAAT':        parse_ziraat,
    'YAPIKREDI':     parse_yapikredi,
    'KUVEYTTURK':    parse_kuveytturk,
    'TURKIYEFINANS': parse_turkiyefinans,
}


# =============================================================================
# 10. Validation
# =============================================================================

def validate_balance_chain(rows: list[TransactionRow]) -> list[BalanceIssue]:
    issues: list[BalanceIssue] = []
    for i in range(1, len(rows)):
        expected = rows[i - 1].balance + rows[i].amount
        if expected != rows[i].balance:
            issues.append(BalanceIssue(
                row_index=i, date=rows[i].date,
                expected=expected, actual=rows[i].balance,
                diff=rows[i].balance - expected,
            ))
    return issues


def validate_final_balance(result: StatementResult) -> Optional[bool]:
    if not result.rows or result.metadata.stated_balance is None:
        return None
    return result.rows[-1].balance == result.metadata.stated_balance


# =============================================================================
# 11. Pipeline
# =============================================================================

def process_pdf(pdf_bytes: bytes, filename: str) -> StatementResult:
    """Main entry point. Wraps every stage in try/except so a single bad
    PDF never takes down the run. If the detected-bank parser returns zero
    rows we automatically retry with the generic block-accumulation parser
    — that way we never silently drop data when a layout drifts."""
    try:
        text = extract_pdf_text(pdf_bytes)
    except Exception as exc:
        result = StatementResult(
            source_filename=filename,
            metadata=StatementMetadata(bank='UNKNOWN'),
            parser_warnings=[f"Could not read PDF text: {exc}"],
        )
        return result

    # Detection: try the embedded text layer first. If it's empty (scanned PDF),
    # fall back to a small OCR sample of page 1 just for bank detection.
    bank = None
    try:
        bank = detect_bank(text) if text.strip() else None
        if bank is None and not text.strip() and OCR_AVAILABLE:
            try:
                sample = _ocr_pages(pdf_bytes, dpi=300, lang='tur')
                if sample:
                    bank = detect_bank(sample[0]['text'])
            except Exception:
                pass
    except Exception as exc:
        bank = None

    # Primary parser attempt.
    result: Optional[StatementResult] = None
    try:
        if bank is not None and bank in PARSERS:
            result = PARSERS[bank](pdf_bytes, text, filename)
        elif bank is not None:
            result = parse_generic(pdf_bytes, text, filename)
            result.metadata.bank = bank
            result.parser_warnings.insert(
                0, f"Detected {bank} but no dedicated parser. Using generic fallback."
            )
        else:
            result = parse_generic(pdf_bytes, text, filename)
            if not text.strip():
                result.parser_warnings.insert(
                    0,
                    "PDF has no embedded text (scanned/image-only) and OCR was "
                    "not available or did not detect a known bank. Install OCR "
                    "dependencies (see top of file) or upload a vector PDF."
                )
            else:
                result.parser_warnings.insert(
                    0, f"[DEBUG] Bank not detected. First 300 chars: {repr(text[:300])}"
                )
    except Exception as exc:
        result = StatementResult(
            source_filename=filename,
            metadata=StatementMetadata(bank=bank or 'UNKNOWN'),
            parser_warnings=[
                f"Dedicated parser for {bank or 'UNKNOWN'} crashed: {exc}. "
                "Falling back to generic parser."
            ],
        )

    # Safety net: if the dedicated parser found zero rows but the PDF has text,
    # retry with the generic block parser. The user gets *something* rather
    # than a silent empty result.
    if result is not None and len(result.rows) == 0 and text.strip() and bank in PARSERS:
        try:
            fallback = parse_generic(pdf_bytes, text, filename)
            if len(fallback.rows) > 0:
                # Keep original metadata (it was bank-specific) but use fallback rows.
                result.rows = fallback.rows
                result.parser_warnings.insert(
                    0,
                    f"Dedicated {bank} parser found 0 rows; recovered "
                    f"{len(fallback.rows)} via generic block parser. "
                    f"Verify the balance chain below before trusting."
                )
        except Exception:
            pass  # Generic fallback also failed — keep the empty result.

    if result is None:
        result = StatementResult(
            source_filename=filename,
            metadata=StatementMetadata(bank='UNKNOWN'),
            parser_warnings=["Unknown error during parsing — no result produced."],
        )

    # Always run validations — even on empty rows, it just yields [].
    try:
        result.balance_issues = validate_balance_chain(result.rows)
        result.final_balance_match = validate_final_balance(result)
    except Exception as exc:
        result.parser_warnings.append(f"Balance validation crashed: {exc}")
    return result


# =============================================================================
# 12. Output — DataFrame & file generation with proper data types
# =============================================================================

OUTPUT_COLUMNS = ['TARİH', 'Saat', 'TUTAR', 'Bakiye', 'AÇIKLAMA', 'DEKONT']


def result_to_dataframe(result: StatementResult) -> pd.DataFrame:
    """Build a DataFrame with proper dtypes:
    - TARİH:    datetime64[ns]   (real dates, not strings)
    - TUTAR:    float64          (monetary number)
    - Bakiye:   float64          (monetary number)
    - DEKONT:   string           (always text — display compatibility)
    - Saat / AÇIKLAMA: string

    DEKONT is always a string here so concatenated DataFrames from different
    banks (where one is all-digit, another alphanumeric) display cleanly
    through pandas/PyArrow. The Excel writer (df_to_xlsx_bytes) decides
    per-export whether to *promote* DEKONT to Int64 in the workbook.
    """
    if not result.rows:
        return pd.DataFrame({
            'TARİH':    pd.Series(dtype='datetime64[ns]'),
            'Saat':     pd.Series(dtype='string'),
            'TUTAR':    pd.Series(dtype='float64'),
            'Bakiye':   pd.Series(dtype='float64'),
            'AÇIKLAMA': pd.Series(dtype='string'),
            'DEKONT':   pd.Series(dtype='string'),
        })[OUTPUT_COLUMNS]

    df = pd.DataFrame({
        'TARİH':    [r.date for r in result.rows],
        'Saat':     [r.time for r in result.rows],
        'TUTAR':    [float(r.amount) for r in result.rows],
        'Bakiye':   [float(r.balance) for r in result.rows],
        'AÇIKLAMA': [r.description for r in result.rows],
        'DEKONT':   [r.receipt for r in result.rows],
    })[OUTPUT_COLUMNS]

    df['TARİH'] = pd.to_datetime(df['TARİH'], format='%d-%m-%Y', errors='coerce')
    df['Saat'] = df['Saat'].astype('string').fillna('')
    df['TUTAR'] = pd.to_numeric(df['TUTAR'], errors='coerce').astype('float64')
    df['Bakiye'] = pd.to_numeric(df['Bakiye'], errors='coerce').astype('float64')
    df['AÇIKLAMA'] = df['AÇIKLAMA'].astype('string').fillna('')
    df['DEKONT'] = df['DEKONT'].astype('string').fillna('')
    return df


def _inject_ignored_errors(xlsx_bytes: bytes, sqref: str) -> bytes:
    """Inject an `<ignoredErrors>` XML block into the saved xlsx so Excel
    suppresses the green-triangle 'Number stored as text' warning over `sqref`.

    Used when DEKONT is a text column containing numeric-looking receipts
    (e.g. 10-digit Halkbank dekonts) — the data must remain text but the
    warning is noise."""
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), 'r') as zin:
        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                content = zin.read(item)
                if item == 'xl/worksheets/sheet1.xml':
                    txt = content.decode('utf-8')
                    block = (
                        '<ignoredErrors>'
                        f'<ignoredError sqref="{sqref}" numberStoredAsText="1"/>'
                        '</ignoredErrors>'
                    )
                    # ignoredErrors must come before extLst / after mergeCells per schema;
                    # placing it just before </worksheet> works for Excel in practice.
                    if '<ignoredErrors' not in txt:
                        txt = txt.replace('</worksheet>', block + '</worksheet>')
                    content = txt.encode('utf-8')
                zout.writestr(item, content)
    return out.getvalue()


def df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Write Excel with proper data types so Excel sees numbers as numbers
    and dates as dates — no green triangles."""
    from openpyxl.utils import get_column_letter

    df = df.copy()

    # DEKONT comes in as `string` dtype (uniform across files). Decide here,
    # at export time, whether to promote it to Int64 (when every non-empty
    # value is purely digits — Halkbank-style 10-digit receipts) so Excel
    # treats it as a number with no green-triangle warning. Otherwise keep
    # it as text and inject `<ignoredErrors>` only if any cell looks numeric.
    dekont_is_numeric = False
    if 'DEKONT' in df.columns and len(df):
        non_empty = df['DEKONT'].astype('string').fillna('').str.strip()
        non_empty = non_empty[non_empty != '']
        if len(non_empty) > 0 and non_empty.str.fullmatch(r'\d+').all():
            df['DEKONT'] = pd.to_numeric(df['DEKONT'], errors='coerce').astype('Int64')
            dekont_is_numeric = True

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Hareketler', index=False)
        ws = writer.sheets['Hareketler']

        for i, col in enumerate(df.columns, start=1):
            try:
                sample = df[col].astype(str)
                width = min(60, max(12, int(sample.str.len().quantile(0.95)) + 2))
            except Exception:
                width = 18
            ws.column_dimensions[get_column_letter(i)].width = width

        n_data_rows = len(df)
        last_row = n_data_rows + 1

        if 'TARİH' in df.columns and n_data_rows:
            col_letter = get_column_letter(df.columns.get_loc('TARİH') + 1)
            for row_idx in range(2, last_row + 1):
                ws[f'{col_letter}{row_idx}'].number_format = 'DD.MM.YYYY'

        for col_name in ('TUTAR', 'Bakiye'):
            if col_name in df.columns and n_data_rows:
                col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
                for row_idx in range(2, last_row + 1):
                    ws[f'{col_letter}{row_idx}'].number_format = '#,##0.00;-#,##0.00'

        if dekont_is_numeric and 'DEKONT' in df.columns and n_data_rows:
            col_letter = get_column_letter(df.columns.get_loc('DEKONT') + 1)
            for row_idx in range(2, last_row + 1):
                ws[f'{col_letter}{row_idx}'].number_format = '0'

    xlsx_bytes = buf.getvalue()

    # Text DEKONT containing some numeric-looking values → suppress green triangles.
    if 'DEKONT' in df.columns and not dekont_is_numeric and len(df):
        has_numericish = df['DEKONT'].astype(str).str.fullmatch(r'\d+').any()
        if has_numericish:
            col_letter = get_column_letter(df.columns.get_loc('DEKONT') + 1)
            sqref = f'{col_letter}2:{col_letter}{len(df) + 1}'
            xlsx_bytes = _inject_ignored_errors(xlsx_bytes, sqref)

    return xlsx_bytes


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    df = df.copy()
    # Format date column as DD-MM-YYYY for CSV (universal & locale-safe)
    if 'TARİH' in df.columns:
        df['TARİH'] = pd.to_datetime(df['TARİH'], errors='coerce').dt.strftime('%d-%m-%Y')
    for col in ('TUTAR', 'Bakiye'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df.to_csv(index=False, float_format='%.2f').encode('utf-8-sig')


# =============================================================================
# 13. Filename
# =============================================================================

_SAFE_RE = re.compile(r'[^A-Za-z0-9._\-]+')


def safe_filename(name: str, max_len: int = 80) -> str:
    cleaned = _SAFE_RE.sub('_', name).strip('_')
    return cleaned[:max_len] if cleaned else 'output'


def build_output_basename(result: StatementResult) -> str:
    pdf_base = os.path.splitext(os.path.basename(result.source_filename))[0]
    cust = result.metadata.customer_name
    if cust:
        cust_short = ' '.join(cust.split()[:3])
        return safe_filename(f"{pdf_base}__{cust_short}")
    return safe_filename(pdf_base)


# =============================================================================
# 14. Upload handling (PDF + ZIP)
# =============================================================================

def iter_uploaded(uploaded_files) -> list[tuple[str, bytes]]:
    out: list[tuple[str, bytes]] = []
    for uf in uploaded_files:
        name = uf.name
        data = uf.read()
        lname = name.lower()
        if lname.endswith('.pdf'):
            out.append((name, data))
        elif lname.endswith('.zip'):
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    for inner in z.namelist():
                        if inner.lower().endswith('.pdf') and not inner.endswith('/'):
                            out.append((os.path.basename(inner), z.read(inner)))
            except zipfile.BadZipFile:
                st.error(f"Could not read {name} as ZIP.")
        else:
            st.warning(f"Skipping unsupported file: {name}")
    return out


# =============================================================================
# 15. Streamlit UI
# =============================================================================

def render_metadata(md: StatementMetadata) -> None:
    cols = st.columns(3)
    cols[0].markdown(f"**Bank**: {md.bank or '—'}")
    cols[0].markdown(f"**Müşteri**: {md.customer_name or '—'}")
    cols[1].markdown(f"**Hesap No**: {md.account_no or '—'}")
    cols[1].markdown(f"**IBAN**: {md.iban or '—'}")
    cols[2].markdown(f"**Dönem**: {md.period_start or '—'} → {md.period_end or '—'}")
    cols[2].markdown(f"**Şube**: {md.branch or '—'}")


def render_validation(result: StatementResult) -> None:
    n_rows = len(result.rows)
    issues = result.balance_issues
    final_match = result.final_balance_match

    if n_rows == 0:
        st.error("❌ No transactions parsed.")
        for w in result.parser_warnings:
            st.warning(w)
        for info in result.parser_info:
            st.info(info)
        return

    if not issues:
        st.success(
            f"✅ Balance chain verified across **{n_rows-1}** consecutive row pairs. "
            f"All {n_rows} transactions reconcile mathematically."
        )
    else:
        st.error(
            f"❌ Balance mismatches on {len(issues)} row(s). "
            f"DO NOT TRUST this output — extraction is broken for at least one row."
        )
        with st.expander(f"Show first {min(10, len(issues))} balance issues"):
            for iss in issues[:10]:
                st.write(
                    f"Row {iss.row_index} ({iss.date}): "
                    f"expected {format_tr_decimal(iss.expected)}, "
                    f"actual {format_tr_decimal(iss.actual)}, "
                    f"diff {format_tr_decimal(iss.diff)}"
                )

    if final_match is True:
        st.success(
            f"✅ Final balance matches statement header: "
            f"**{format_tr_decimal(result.metadata.stated_balance)}**."
        )
    elif final_match is False:
        last = result.rows[-1].balance
        stated = result.metadata.stated_balance
        st.error(
            f"❌ Final balance mismatch — last row says "
            f"**{format_tr_decimal(last)}**, statement header says "
            f"**{format_tr_decimal(stated)}**."
        )

    # Actionable warnings always shown.
    for w in result.parser_warnings:
        st.warning(w)

    # Informational notes (e.g. "OCR was used", repairs that the balance
    # chain already validated) → hidden in an expander when everything
    # reconciled, surfaced inline when there are issues.
    if result.parser_info:
        if issues or final_match is False:
            for info in result.parser_info:
                st.info(info)
        else:
            with st.expander(f"Parser notes ({len(result.parser_info)})"):
                for info in result.parser_info:
                    st.write(f"• {info}")


def main() -> None:
    st.set_page_config(page_title="Bank Statement Extractor", layout="wide")
    st.title("Bank Statement → Excel/CSV")
    st.caption(
        "Accountant-grade extraction for Turkish bank PDFs. "
        "Supported: " + ", ".join(PARSERS.keys()) + ". "
        "Other banks: generic fallback parser."
    )

    if not OCR_AVAILABLE:
        with st.expander("ℹ OCR not available (only needed for scanned PDFs)"):
            st.write(
                "Vector PDFs (Halkbank, Akbank, vector-mode Ziraat) work fine. "
                "Scanned/image-only PDFs won't be readable until OCR is enabled."
            )
            st.code(
                "# Linux / Streamlit Cloud (via packages.txt):\n"
                "tesseract-ocr\ntesseract-ocr-tur\npoppler-utils\n\n"
                "# macOS:\nbrew install tesseract tesseract-lang poppler\n\n"
                "# Windows: install from\n"
                "#   https://github.com/UB-Mannheim/tesseract/wiki\n"
                "#   https://github.com/oschwartz10612/poppler-windows/releases",
                language='text',
            )
            if OCR_IMPORT_ERROR:
                st.caption(f"Detail: {OCR_IMPORT_ERROR}")

    if 'results' not in st.session_state:
        st.session_state.results = []

    uploaded = st.file_uploader(
        "Upload PDF statements or a ZIP folder containing them",
        type=['pdf', 'zip'],
        accept_multiple_files=True,
    )

    col_a, col_b = st.columns([1, 1])
    if col_a.button("Process uploads", type="primary", disabled=not uploaded):
        files = iter_uploaded(uploaded)
        if not files:
            st.warning("No PDFs found in uploads.")
        new_results: list[StatementResult] = []
        progress = st.progress(0.0, text="Processing…")
        for i, (name, data) in enumerate(files, start=1):
            try:
                res = process_pdf(data, name)
            except Exception as exc:
                res = StatementResult(source_filename=name, metadata=StatementMetadata())
                res.parser_warnings.append(f"Fatal error: {exc}")
            new_results.append(res)
            progress.progress(i / len(files), text=f"Processed {i}/{len(files)}")
        progress.empty()
        st.session_state.results.extend(new_results)
        st.success(f"Added {len(new_results)} file(s). Total: "
                   f"{len(st.session_state.results)}.")

    if col_b.button("Clear all", disabled=not st.session_state.results):
        st.session_state.results = []
        st.rerun()

    if not st.session_state.results:
        st.info("No statements processed yet. Upload PDFs above to begin.")
        return

    st.divider()
    st.subheader("Per-file results")
    for idx, res in enumerate(st.session_state.results):
        n_rows = len(res.rows)
        ok = (not res.balance_issues) and (res.final_balance_match is not False) and n_rows > 0
        badge = "✅" if ok else ("⚠️" if n_rows > 0 else "❌")
        with st.expander(f"{badge} {res.source_filename} — {n_rows} rows", expanded=False):
            render_metadata(res.metadata)
            st.markdown("---")
            render_validation(res)

            if n_rows > 0:
                df = result_to_dataframe(res)
                st.dataframe(df.head(20), width='stretch')

                base = build_output_basename(res)
                dl1, dl2 = st.columns(2)
                dl1.download_button(
                    f"⬇ {base}.xlsx",
                    data=df_to_xlsx_bytes(df),
                    file_name=f"{base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"xlsx_{idx}",
                )
                dl2.download_button(
                    f"⬇ {base}.csv",
                    data=df_to_csv_bytes(df),
                    file_name=f"{base}.csv",
                    mime="text/csv",
                    key=f"csv_{idx}",
                )

    st.divider()
    st.subheader("Combined output (all uploaded files)")
    all_dfs = [result_to_dataframe(r) for r in st.session_state.results if r.rows]
    if not all_dfs:
        st.info("No rows yet to combine.")
        return
    combined = pd.concat(all_dfs, ignore_index=True)

    before = len(combined)
    combined = combined.drop_duplicates(
        subset=['TARİH', 'TUTAR', 'Bakiye', 'AÇIKLAMA'],
        keep='first',
    ).reset_index(drop=True)
    after = len(combined)
    if before != after:
        st.info(f"Removed {before - after} duplicate row(s) across uploads.")

    st.write(f"**Total rows: {len(combined):,}**")
    st.dataframe(combined.head(50), width='stretch')

    dl1, dl2 = st.columns(2)
    dl1.download_button(
        "⬇ combined.xlsx",
        data=df_to_xlsx_bytes(combined),
        file_name="combined_bank_statements.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    dl2.download_button(
        "⬇ combined.csv",
        data=df_to_csv_bytes(combined),
        file_name="combined_bank_statements.csv",
        mime="text/csv",
    )


if __name__ == '__main__':
    main()